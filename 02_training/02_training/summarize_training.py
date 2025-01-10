import os

from openpyxl import Workbook
from openpyxl.chart import (
  LineChart,
  Reference,
  Series,
)

this_file_path = os.path.dirname(os.path.realpath(__file__))

bootstraps = 10
initial_weights = 3
epochs = 30

wb = Workbook()
wb.remove(wb.active)

for training_config in [
  # params destination
  # ('binADNI_weights/T1_@MNI_nlin',),
  # ('binADNI_weights/T1_@MNI_nlin_bet',),
  # ('binADNI_weights/T1_@MNI_nlin_bet_bin01375',),
  # ('binADNI_weights/T1_@MNI_nlin_bet_bin',),
  # ('binADNI_weights/T1_@MNI_nlin_bet_bin04125',),
  # ('binADNI_weights/T1_@MNI_nlin_bin01375',),
  # ('binADNI_weights/T1_@MNI_nlin_bin',),
  # ('binADNI_weights/T1_@MNI_nlin_bin04125',),
  # ('binADNI_weights/T1_@MNI_nlin_bet_RG',),
  # ('binADNI_weights/T1_@MNI_nlin_bet_bin_RG',),

  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin_bin01375',),
  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin_bet_bin01375',),

  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin_bin0275',),
  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin_bet_bin0275',),
  
  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin_bin04125',),
  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin_bet_bin04125',),

  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin',),
  ('binADNI_weights_30_epochs__new_normalizer/T1_@MNI_nlin_bet',),
]:
  (params_destination,) = training_config
  is_RG = params_destination.endswith('_RG') or params_destination.endswith('_RG_MNI') or params_destination.endswith('_RG_BG')

  params_destination_full_path = os.path.join(this_file_path, params_destination)
  if not os.path.exists(params_destination_full_path):
    continue

  param_names = os.listdir(params_destination_full_path)
  param_names.sort()

  ws = wb.create_sheet(params_destination.split('/')[-1])

  row = [
    'bootstrap_index', 'initial_weights_index', 'epoch',
    'training_classification_loss', 'validation_classification_loss',
    'training_accuracy', 'validation_accuracy',
  ]
  if is_RG:
    row = row + [
      'training_relevance_inner_mask', 'validation_relevance_inner_mask',
    ]
  ws.append(row)

  # T1__boostrap_index-01__initial_weights_index-03__055-tcl-0.336-vcl-0.505-tca-0.845-vca-0.778-tsrim-0.991-vsrim-0.992.h5
  for param_name in param_names:
    row = [
      int(param_name[len('T1__bootstrap_index-'):len('T1__bootstrap_index-') + 2]),
      int(param_name[len('T1__bootstrap_index-01__initial_weights_index-'):len('T1__bootstrap_index-01__initial_weights_index-') + 2]),
      int(param_name[len('T1__bootstrap_index-01__initial_weights_index-01__'):len('T1__bootstrap_index-01__initial_weights_index-01__') + 3]),
      .0, # float(param_name[param_name.index('tcl-') + len('tcl-'):param_name.index('vcl-') - 1]),
      .0, # float(param_name[param_name.index('vcl-') + len('vcl-'):param_name.index('tca-') - 1]),
      float(param_name[param_name.index('tca-') + len('tca-'):param_name.index('vca-') - 1]),
      float(param_name[param_name.index('vca-') + len('vca-'):param_name.index('tim--') - 1 if is_RG else param_name.index('.pth')]),
    ]
    if is_RG:
      # T1__bootstrap_index-01__initial_weights_index-01__001-tcl--0.256-vcl--0.350-tca-0.587-vca-0.656-tim--0.923-vim--0.981.pth
      row = row + [
        float(param_name[param_name.index('tim--') + len('tim--'):param_name.index('vim--') - 1]),
        float(param_name[param_name.index('vim--') + len('vim--'):param_name.index('.pth')]),
      ]
    
    if row[2] <= epochs:
      ws.append(row)

  ws['K1'] = 'bootstrap index'
  ws['L1'] = 'initial weights index'
  ws['M1'] = 'max. val. acc.'
  ws['N1'] = 'relative index (epoch)'
  ws['O1'] = 'last val. acc.'
  ws['P1'] = 'delta val. acc.'
  for bootstrap_index in range(bootstraps):
    for initial_weights_index in range(initial_weights):
      row_number = bootstrap_index * initial_weights + initial_weights_index + 2
      ws['K' + str(row_number)] = bootstrap_index + 1
      ws['L' + str(row_number)] = initial_weights_index + 1
      ws['M' + str(row_number)] = f'=MAX(G{bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + 2}:G{bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + epochs + 1})'
      ws['N' + str(row_number)] = f'=MATCH(MAX(G{bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + 2}:G{bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + epochs + 1}), G{bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + 2}:G{bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + epochs + 1}, 0)'
      ws['O' + str(row_number)] = f'=G{bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + epochs + 1}'
      ws['P' + str(row_number)] = f'=M{row_number}-O{row_number}'
 
  # ws['S8'] = 'Max class. acc. val. overall'
  # ws['S9'] = 'Avg class. acc. val. overall'
  # ws['T8'] = '=MAX(T2:T6)'
  # ws['T9'] = '=AVERAGE(T2:T6)'
  # ws['S11'] = 'Indices'
  # ws['T11'] = '="["&U2&", "&U3&", "&U4&", "&U5&", "&U6&"]"'

  measures = [
    'cat. cross entropy training',
    'cat. cross entropy validation',
    'class. accuracy training',
    'class. accuracy validation',
  ]
  if is_RG:
    measures = measures + [
      'relevance inner mask training',
      'relevance inner mask validation',
    ]

  for measure_index, measure in enumerate(measures):
    chart = LineChart()
    chart.title = measure
    chart.height = 15
    chart.width = 30
    chart.x_axis.title = 'epoch'
    chart.x_axis.delete = False # for LibreOffice
    chart.y_axis.title = 'value'
    chart.y_axis.delete = False # for LibreOffice
    
    for bootstrap_index in range(bootstraps):
      for initial_weights_index in range(initial_weights):
        y = Reference(ws, min_col=4 + measure_index, min_row=bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + 2, max_row=bootstrap_index * initial_weights * epochs + initial_weights_index * epochs + epochs + 1)
        s = Series(y, title='bs ' + str(bootstrap_index + 1) + ' iw '+ str(initial_weights_index + 1))
        chart.append(s)
    
    chart.set_categories(Reference(ws, min_col=3, min_row=2, max_row=epochs + 1))

    ws.add_chart(chart, 'K' + str(32 + measure_index * 30))

wb.save('training_summary_30_epochs__new_normalizer.xlsx')
