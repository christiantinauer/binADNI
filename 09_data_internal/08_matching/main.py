import json
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import functools

with open('../01_gather/02_CN.json', 'r') as infile:
  ASPSF = json.load(infile)

with open('../03_images_sanity/ASPSF_excludes_preprocess_crashes.json', 'r') as infile:
  ASPSF_excludes_preprocess_crashes = json.load(infile)

with open('../03_images_sanity/ASPSF_excludes_image_sanity_checks.json', 'r') as infile:
  ASPSF_excludes_image_sanity_checks = json.load(infile)

with open('../01_gather/02_AD.json', 'r') as infile:
  ProDem = json.load(infile)

with open('../03_images_sanity/ProDem_excludes_preprocess_crashes.json', 'r') as infile:
  ProDem_excludes_preprocess_crashes = json.load(infile)

with open('../03_images_sanity/ProDem_excludes_image_sanity_checks.json', 'r') as infile:
  ProDem_excludes_image_sanity_checks = json.load(infile)

ASPSF_age_matching = list(set(ASPSF) - set(ASPSF_excludes_preprocess_crashes) - set(ASPSF_excludes_image_sanity_checks))
ASPSF_age_matching.sort()

ProDem_age_matching = list(set(ProDem) - set(ProDem_excludes_preprocess_crashes) - set(ProDem_excludes_image_sanity_checks))
ProDem_age_matching.sort()

ASPSF_workbook = load_workbook('ASPS_age_sex_atScandate.xlsx', data_only=True)
ProDem_workbook = load_workbook('ProDem_ageAtScandate_sex_updated2018_11_05.xlsx', data_only=True)

ASPSF_worksheet = ASPSF_workbook['ASPSF']
ProDem_worksheet = ProDem_workbook['SCANID']

# CN matched age 95% quantile: 74.7/[62.235 86.1  ]
# AD matched age 95% quantile: 76.0/[58.3 89.7]

CN_age_lower_bound = 55.7# 58.3 # 62.235
CN_age_upper_bound = 100# 89.7 # 86.1
AD_age_lower_bound = 55.7# 58.3
AD_age_upper_bound = 100# 89.7

image_ids_to_features = {}

# load age data from xlsx
ASPSF_data = {}
for row_index in range(512):
  subject_id = ASPSF_worksheet.cell(row=row_index + 1, column=1).value
  if subject_id not in ASPSF_age_matching:
    continue
  
  subject_age = float(ASPSF_worksheet.cell(row=row_index + 1, column=3).value)
  subject_sex = 1 if str(ASPSF_worksheet.cell(row=row_index + 1, column=2).value) == 'M' else 2

  if subject_age > CN_age_lower_bound and subject_age < CN_age_upper_bound:
    ASPSF_data[subject_id] = { 'age': subject_age }

    image_ids_to_features[subject_id] = {
			'age': subject_age,
      'sex': subject_sex,
      'subject_id': subject_id[0:len('803055')]
		}

ProDem_data = {}
for row_index in range(1109):
  subject_id = ProDem_worksheet.cell(row=row_index + 1, column=1).value
  # 801 are scans from Graz, 803 are scans from Vienna, both centers used 3T
  if not subject_id.startswith('801') and not subject_id.startswith('803') or subject_id not in ProDem_age_matching:
    continue

  subject_age = float(ProDem_worksheet.cell(row=row_index + 1, column=4).value)
  subject_sex = int(ProDem_worksheet.cell(row=row_index + 1, column=5).value)

  if subject_age > AD_age_lower_bound and subject_age < AD_age_upper_bound:
    ProDem_data[subject_id] = { 'age': subject_age }

    image_ids_to_features[subject_id] = {
			'age': subject_age,
      'sex': subject_sex,
      'subject_id': subject_id[0:len('803055')]
		}

ASPSF_ids = list(ASPSF_data.keys())
ASPSF_ids.sort()

with open('ASPSF_matched.json', 'w') as outfile:
  json.dump(ASPSF_ids, outfile, indent=2)

ProDem_ids = list(ProDem_data.keys())
ProDem_ids.sort()

with open('ProDem_matched.json', 'w') as outfile:
  json.dump(ProDem_ids, outfile, indent=2)

CN_matched_ages = list(map(lambda e: image_ids_to_features[e]['age'], ASPSF_ids))
AD_matched_ages = list(map(lambda e: image_ids_to_features[e]['age'], ProDem_ids))

plt.hist(CN_matched_ages, bins='fd', density=False, alpha=.5)
plt.hist(AD_matched_ages, bins='fd', density=False, alpha=.5)
plt.legend(['CN', 'AD'])
plt.xlabel('age')
plt.ylabel('count')
plt.savefig('age_hists.png')
plt.clf()

print('CN matched age min/max: ' + str(np.min(CN_matched_ages)) + '/' + str(np.max(CN_matched_ages)))
print('AD matched age min/max: ' + str(np.min(AD_matched_ages)) + '/' + str(np.max(AD_matched_ages)))

print('CN matched age median/IQR: ' + str(np.median(CN_matched_ages)) + '/' + str(np.percentile(CN_matched_ages, (25, 75))))
print('AD matched age median/IQR: ' + str(np.median(AD_matched_ages)) + '/' + str(np.percentile(AD_matched_ages, (25, 75))))

print('CN matched age 95% quantile: ' + str(np.median(CN_matched_ages)) + '/' + str(np.percentile(CN_matched_ages, (2.5, 97.5))))
print('AD matched age 95% quantile: ' + str(np.median(AD_matched_ages)) + '/' + str(np.percentile(AD_matched_ages, (2.5, 97.5))))

print('CN matched age mean/std: ' + str(np.mean(CN_matched_ages)) + '/' + str(np.std(CN_matched_ages)))
print('AD matched age mean/std: ' + str(np.mean(AD_matched_ages)) + '/' + str(np.std(AD_matched_ages)))

CN_matched_sexes = np.array(list(map(lambda e: image_ids_to_features[e]['sex'], ASPSF_ids)))
AD_matched_sexes = np.array(list(map(lambda e: image_ids_to_features[e]['sex'], ProDem_ids)))

print('CN matched sex m/f: ' + str(np.count_nonzero(CN_matched_sexes == 1)) + '/' + str(np.count_nonzero(CN_matched_sexes == 2)))
print('AD matched sex m/f: ' + str(np.count_nonzero(AD_matched_sexes == 1)) + '/' + str(np.count_nonzero(AD_matched_sexes == 2)))

CN_matched_subjects = list(set(map(lambda e: image_ids_to_features[e]['subject_id'], ASPSF_ids)))
AD_matched_subjects = list(set(map(lambda e: image_ids_to_features[e]['subject_id'], ProDem_ids)))

print('CN matched subjects/images: ' + str(len(CN_matched_subjects)) + '/' + str(len(ASPSF_ids)))
print('AD matched subjects/images: ' + str(len(AD_matched_subjects)) + '/' + str(len(ProDem_ids)))

def reducer_sex(a, s):
  a[image_ids_to_features[s]['subject_id']] = image_ids_to_features[s]['sex']
  return a

CN_matched_subject_sexes = np.array(list(functools.reduce(reducer_sex, ASPSF_ids, {}).values()))
AD_matched_subject_sexes = np.array(list(functools.reduce(reducer_sex, ProDem_ids, {}).values()))

print('CN subjects matched m/f: ' + str(np.count_nonzero(CN_matched_subject_sexes == 1)) + '/' + str(np.count_nonzero(CN_matched_subject_sexes == 2)))
print('AD subjects matched m/f: ' + str(np.count_nonzero(AD_matched_subject_sexes == 1)) + '/' + str(np.count_nonzero(AD_matched_subject_sexes == 2)))

APSPF_MMSE_workbook = load_workbook('ASPSFam_MMSE_CDR.xlsx', data_only=True)
ASPSF_MMSE_worksheet = APSPF_MMSE_workbook['ASPSFam_MMSE_CDR']

ASPSF_ids_baseline = set(id[0:6] for id in ASPSF_ids)
ASPSF_mmse_scores = []
ASPSF_cdr_scores = []

for row_index in range(1, 421):
  subject_id = f'{int(ASPSF_MMSE_worksheet.cell(row=row_index + 1, column=1).value):06d}'
  mmse_score = ASPSF_MMSE_worksheet.cell(row=row_index + 1, column=3).value
  cdr_score  = str(ASPSF_MMSE_worksheet.cell(row=row_index + 1, column=6).value or '')
  
  if subject_id in ASPSF_ids_baseline:
    if mmse_score is not None:
      ASPSF_mmse_scores.append(int(mmse_score))
    
    if cdr_score != '':
      ASPSF_cdr_scores.append(f'{float(cdr_score):1.1f}')

print('CN matched mmse mean/std: ' + str(np.mean(ASPSF_mmse_scores)) + '/' + str(np.std(ASPSF_mmse_scores)))
print('CN matched mmse min/max: ' + str(np.min(ASPSF_mmse_scores)) + '/' + str(np.max(ASPSF_mmse_scores)))
print('CN matched mmse n/a: ' + str(len(ASPSF_ids_baseline) - len(ASPSF_mmse_scores)))

ProDem_MMSE_workbook = load_workbook('PRODEM_MMSE_CDR.xlsx', data_only=True)
ProDem_MMSE_worksheet = ProDem_MMSE_workbook['PRODEM_MMSE_CDR']

ProDem_ids_baseline = set(id[0:6] for id in ProDem_ids)
ProDem_mmse_scores = []
ProDem_cdr_scores = []

for row_index in range(1, 820):
  subject_id = str(ProDem_MMSE_worksheet.cell(row=row_index + 1, column=1).value)
  mmse_score = ProDem_MMSE_worksheet.cell(row=row_index + 1, column=3).value
  cdr_score = str(ProDem_MMSE_worksheet.cell(row=row_index + 1, column=4).value or '')
  
  if subject_id in ProDem_ids_baseline:
    if mmse_score is not None:
      ProDem_mmse_scores.append(int(mmse_score))
    
    if cdr_score != '':
      ProDem_cdr_scores.append(f'{float(cdr_score):1.1f}')

print('AD matched mmse mean/std: ' + str(np.mean(ProDem_mmse_scores)) + '/' + str(np.std(ProDem_mmse_scores)))
print('AD matched mmse min/max: ' + str(np.min(ProDem_mmse_scores)) + '/' + str(np.max(ProDem_mmse_scores)))
print('AD matched mmse n/a: ' + str(len(ProDem_ids_baseline) - len(ProDem_mmse_scores)))

ASPSF_cdr_scores = np.array(ASPSF_cdr_scores)

print('CN subjects matched cdr 0.0/0.5/1.0/1.5/2.0/2.5/3.0/empty: ' +\
	str(np.count_nonzero(ASPSF_cdr_scores == '0.0')) + '/' +\
	str(np.count_nonzero(ASPSF_cdr_scores == '0.5')) + '/' +\
	str(np.count_nonzero(ASPSF_cdr_scores == '1.0')) + '/' +\
	str(np.count_nonzero(ASPSF_cdr_scores == '1.5')) + '/' +\
	str(np.count_nonzero(ASPSF_cdr_scores == '2.0')) + '/' +\
  str(np.count_nonzero(ASPSF_cdr_scores == '2.5')) + '/' +\
  str(np.count_nonzero(ASPSF_cdr_scores == '3.0')) + '/' +\
  str(len(ASPSF_ids_baseline) - len(ASPSF_cdr_scores)))

ProDem_cdr_scores = np.array(ProDem_cdr_scores)

print('AD subjects matched cdr 0.0/0.5/1.0/1.5/2.0/2.5/3.0/empty: ' +\
	str(np.count_nonzero(ProDem_cdr_scores == '0.0')) + '/' +\
	str(np.count_nonzero(ProDem_cdr_scores == '0.5')) + '/' +\
	str(np.count_nonzero(ProDem_cdr_scores == '1.0')) + '/' +\
	str(np.count_nonzero(ProDem_cdr_scores == '1.5')) + '/' +\
	str(np.count_nonzero(ProDem_cdr_scores == '2.0')) + '/' +\
  str(np.count_nonzero(ProDem_cdr_scores == '2.5')) + '/' +\
  str(np.count_nonzero(ProDem_cdr_scores == '3.0')) + '/' +\
  str(len(ProDem_ids_baseline) - len(ProDem_cdr_scores)))

# APOE and Education
APSPF_apoe4_workbook = load_workbook('ASPSFam_edu_ApoE.xlsx', data_only=True)
ASPSF_apoe4_worksheet = APSPF_apoe4_workbook['ASPSFam_edu_ApoE']

ASPSF_education_in_years = []
ASPSF_apoe = []

for row_index in range(1, 421):
  subject_id = f'{int(ASPSF_apoe4_worksheet.cell(row=row_index + 1, column=1).value):06d}'
  education = ASPSF_apoe4_worksheet.cell(row=row_index + 1, column=2).value
  apoe = str(ASPSF_apoe4_worksheet.cell(row=row_index + 1, column=3).value) + '/' + str(ASPSF_apoe4_worksheet.cell(row=row_index + 1, column=4).value)
  
  # 1 = Grundschule (Volks + Hauptschule) = 9 Jahre
  # 2 = Lehre = 10 Jahre
  # 3 = AHS / BHS /LBA (LBA =Lehrerbildungsanstalt) = 13 Jahre
  # 4 = Hochschule = 18 Jahre
  education_in_years = None
  if education == 1:
    education_in_years = 9
  elif education == 2:
    education_in_years = 10
  elif education == 3:
    education_in_years = 13
  elif education == 4:
    education_in_years = 18

  if subject_id in ASPSF_ids_baseline:
    if education_in_years is not None:
      ASPSF_education_in_years.append(education_in_years)
    
    ASPSF_apoe.append(apoe)

print('CN matched education in years mean/std: ' + str(np.mean(ASPSF_education_in_years)) + '/' + str(np.std(ASPSF_education_in_years)))
print('CN matched education in years min/max: ' + str(np.min(ASPSF_education_in_years)) + '/' + str(np.max(ASPSF_education_in_years)))
print('CN matched education in years n/a: ' + str(len(ASPSF_ids_baseline) - len(ASPSF_education_in_years)))

ASPSF_apoe = np.array(ASPSF_apoe)
# print(ASPSF_apoe)

print('CN subjects matched APOE ε2/ε2 ε2/ε3 ε2/ε4 ε3/ε3 ε3/ε4 ε4/ε4 empty: ' +\
	str(np.count_nonzero(ASPSF_apoe == '2/2')) + ' ' +\
	str(np.count_nonzero(ASPSF_apoe == '2/3')) + ' ' +\
	str(np.count_nonzero(ASPSF_apoe == '2/4')) + ' ' +\
	str(np.count_nonzero(ASPSF_apoe == '3/3')) + ' ' +\
	str(np.count_nonzero(ASPSF_apoe == '3/4')) + ' ' +\
	str(np.count_nonzero(ASPSF_apoe == '4/4')) + ' ' +\
	str(np.count_nonzero(ASPSF_apoe == '#NULL!/#NULL!')))

ProDem_apoe_workbook = load_workbook('PRODEM_edu_apoe.xlsx', data_only=True)
ProDem_apoe_worksheet = ProDem_apoe_workbook['PRODEM_edu_apoe']

ProDem_education_in_years = []
ProDem_apoe = []

for row_index in range(1, 853):
  subject_id = str(ProDem_apoe_worksheet.cell(row=row_index + 1, column=1).value)
  education = ProDem_apoe_worksheet.cell(row=row_index + 1, column=2).value
  apoe = str(ProDem_apoe_worksheet.cell(row=row_index + 1, column=3).value) + '/' + str(ProDem_apoe_worksheet.cell(row=row_index + 1, column=4).value)

  # 1 = Gundschule (Volks + Hauptschule) = 9 Jahre
  # 2 =  Lehre =  10 Jahre
  # 3 = AHS = 13 Jahre
  # 4 = BHS = 13 Jahre
  # 5 = LBA (Lehrerbildungsanstalt) = 13 Jahre
  # 6 = Hochschule = 18 Jahre
  education_in_years = None
  if education == 1:
    education_in_years = 9
  elif education == 2:
    education_in_years = 10
  elif education == 3 or education == 4 or education == 5:
    education_in_years = 13
  elif education == 6:
    education_in_years = 18

  if subject_id in ProDem_ids_baseline:
    if education_in_years is not None:
      ProDem_education_in_years.append(education_in_years)
    
    ProDem_apoe.append(apoe)

print('AD matched education in years mean/std: ' + str(np.mean(ProDem_education_in_years)) + '/' + str(np.std(ProDem_education_in_years)))
print('AD matched education in years min/max: ' + str(np.min(ProDem_education_in_years)) + '/' + str(np.max(ProDem_education_in_years)))
print('AD matched education in years n/a: ' + str(len(ProDem_ids_baseline) - len(ProDem_education_in_years)))

ProDem_apoe = np.array(ProDem_apoe)
# print(ProDem_apoe)

print('AD subjects matched APOE ε2/ε2 ε2/ε3 ε2/ε4 ε3/ε3 ε3/ε4 ε4/ε4 empty: ' +\
	str(np.count_nonzero(ProDem_apoe == '2/2')) + ' ' +\
	str(np.count_nonzero(ProDem_apoe == '2/3')) + ' ' +\
	str(np.count_nonzero(ProDem_apoe == '2/4')) + ' ' +\
	str(np.count_nonzero(ProDem_apoe == '3/3')) + ' ' +\
	str(np.count_nonzero(ProDem_apoe == '3/4')) + ' ' +\
	str(np.count_nonzero(ProDem_apoe == '4/4')) + ' ' +\
	str(np.count_nonzero(ProDem_apoe == '#NULL!/#NULL!')))
