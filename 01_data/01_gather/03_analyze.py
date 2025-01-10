import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

details = ['Image Id', 'Acquisition Plane', 'Acquisition Type', 'Coil', 'Field Strength', 'Flip Angle', 'Manufacturer', 'Matrix X', 'Matrix Y', 'Matrix Z', 'Mfg Model', 'Pixel Spacing X', 'Pixel Spacing Y', 'Pulse Sequence', 'Slice Thickness', 'TE', 'TI', 'TR', 'Weighting']
with open('02_idaSearch_11_30_2022_with_details.csv', 'r') as csvfile:
	reader = csv.reader(csvfile, delimiter=',', quotechar='"')
	# skip header
	row = next(reader, None)
	if row is not None:
		ws.append(row[:19] + details)
  
	for row in reader:
		image_id = row[20]
		image_protocol_details = row[21]

		row_imaging_details = [image_id, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]

		for detail in image_protocol_details.split(';'):
			detail = detail.strip()
			if not detail:
				continue

			(name, value) = detail.split('=')
			name = name.strip()
			value = value.strip()

			row_imaging_details[details.index(name)] = value

		ws.append(row[:19] + row_imaging_details)

ws.auto_filter.ref = "A1:" + get_column_letter(ws.max_column) + str(ws.max_row)
wb.save('03_ADNI_details.xlsx')
